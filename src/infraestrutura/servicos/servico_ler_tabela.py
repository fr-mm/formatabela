from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from src.dominio.tabela.interfaces import InterfaceServicoLerTabela
from src.dominio.tabela.otds import OTDTabela, OTDCelula
from src.infraestrutura.enums.enum_servico_ler_tabela import EnumServicoLerTabela


class ServicoLerTabela(InterfaceServicoLerTabela):
    def executar(self, arquivo: str) -> OTDTabela:
        tabela = self.__ler_tabela(arquivo)
        pagina = self.__extrair_primeira_pagina(tabela)
        return OTDTabela(
            nome=self.__extrair_nome(arquivo),
            nome_da_pagina=pagina.title,
            conteudo=self.__extrair_conteudo(pagina)
        )

    @staticmethod
    def __ler_tabela(arquivo: str) -> Workbook:
        return load_workbook(arquivo)

    @staticmethod
    def __extrair_primeira_pagina(tabela: Workbook) -> Worksheet:
        return tabela.worksheets[0]

    @staticmethod
    def __extrair_nome(arquivo: str) -> str:
        sem_caminho = arquivo.split('/')[-1]
        sem_extensao = sem_caminho.split('.')[0]
        return sem_extensao

    def __extrair_conteudo(self, pagina: Worksheet) -> List[List[OTDCelula]]:
        contagem_linhas_preenchidas = self.__contar_conjunto_com_primeira_celula_preenchida(pagina.rows)
        contagem_colunas_preenchidas = self.__contar_conjunto_com_primeira_celula_preenchida(pagina.columns)

        conteudo = []
        for indice_linha in self.__range(contagem_linhas_preenchidas):
            linha = []
            for indice_coluna in self.__range(contagem_colunas_preenchidas):
                celula = pagina.cell(row=indice_linha, column=indice_coluna)
                otd_celula_entrada = self.__criar_otd_celula_entrada(celula)
                linha.append(otd_celula_entrada)
            conteudo.append(linha)
        return conteudo

    @staticmethod
    def __contar_conjunto_com_primeira_celula_preenchida(conjunto: Tuple) -> int:
        return len([celula for celula in conjunto if celula[0].value is not None])

    @staticmethod
    def __range(maximo: int) -> range:
        return range(1, maximo + 1)

    def __criar_otd_celula_entrada(self, celula: Cell) -> OTDCelula:
        return OTDCelula(
            conteudo=self.__extrair_conteudo_da_celula(celula),
            comentario=self.__formatar_comentario(celula.comment),
            cor_fundo=self.__extrair_cor_do_fundo(celula),
            cor_texto=self.__extrair_cor_do_texto(celula)
        )

    @staticmethod
    def __extrair_conteudo_da_celula(celula: Cell) -> str:
        return celula.value or EnumServicoLerTabela.VALOR_PADRAO.value

    @staticmethod
    def __formatar_comentario(comentario: Comment) -> str:
        try:
            formatado = comentario.text
        except AttributeError:
            formatado = EnumServicoLerTabela.VALOR_PADRAO.value
        return formatado

    @staticmethod
    def __extrair_cor_do_texto(celula: Cell) -> str:
        try:
            cor = celula.font.color.rgb
            if cor == EnumServicoLerTabela.COR_TEXT_PADRAO.value:
                raise AttributeError
        except AttributeError:
            cor = EnumServicoLerTabela.VALOR_PADRAO.value
        return cor

    @staticmethod
    def __extrair_cor_do_fundo(celula: Cell) -> str:
        try:
            cor = celula.fill.start_color.rgb
            if cor == EnumServicoLerTabela.COR_FUNDO_PADRARO.value:
                raise AttributeError
        except AttributeError:
            cor = EnumServicoLerTabela.VALOR_PADRAO.value
        return cor
